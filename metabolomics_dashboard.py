"""
MetaFlow Metabolomics Dashboard
Multi-project dashboard — project selected via URL query param ?project=<id>
"""

import os
import io
import json
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import openpyxl
import re
from collections import Counter
from scipy import stats as _scipy_stats
import warnings
warnings.filterwarnings("ignore")

# ─── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="MetaFlow Dashboard",
    page_icon="🧬",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── Custom CSS ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
    /* ── global text darkening ── */
    div[data-testid="stMarkdownContainer"] p { color: #1a1a1a; }
    .stText, label { color: #1a1a1a; }

    /* ── header: override dark-text rule with higher specificity ── */
    .main-header {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
        padding: 2rem; border-radius: 12px; margin-bottom: 1.5rem;
        text-align: center;
    }
    .main-header h1 { font-size: 2.4rem; margin: 0; font-weight: 700; color: #ffffff !important; }
    div[data-testid="stMarkdownContainer"] .main-header p,
    .main-header p  { margin: 0.4rem 0 0; font-size: 1.1rem; color: #e8eaf6 !important; }

    .metric-card {
        background: #f0f2f6; border-radius: 10px; padding: 1.2rem;
        border-left: 5px solid #0f3460; text-align: center;
    }
    .metric-card .value { font-size: 2rem; font-weight: 700; color: #0f3460; }
    .metric-card .label { font-size: 0.9rem; font-weight: 600; color: #333; margin-top: 0.3rem; }

    .section-header {
        background: #0f3460; color: white; padding: 0.7rem 1.2rem;
        border-radius: 8px; margin: 1.5rem 0 0.8rem;
        font-weight: 700; font-size: 1rem; letter-spacing: 0.02em;
    }
    .badge-adhd { background:#c0392b; color:white; padding:3px 10px; border-radius:12px;
                  font-size:0.85rem; font-weight:600; }
    .badge-ctrl { background:#1a6ea8; color:white; padding:3px 10px; border-radius:12px;
                  font-size:0.85rem; font-weight:600; }

    div[data-testid="stExpander"] {
        border: 1px solid #bbb; border-radius: 8px;
    }
    div[data-testid="stExpander"] summary {
        font-size: 1rem; font-weight: 600; color: #111;
    }
    /* sidebar labels */
    .css-1d391kg, .css-qrbaxs { color: #111 !important; font-weight: 500; }

    /* tab labels */
    button[data-baseweb="tab"] { font-size: 0.95rem; font-weight: 600; }
</style>
""", unsafe_allow_html=True)

# ─── Plot template (dark axis text) ───────────────────────────────────────────
PLOT_TEMPLATE = dict(
    layout=dict(
        font=dict(family="Arial, sans-serif", size=13, color="#111"),
        title=dict(font=dict(size=15, color="#111"), x=0.02),
        xaxis=dict(tickfont=dict(size=12, color="#222"), title_font=dict(size=13, color="#222"),
                   linecolor="#888", gridcolor="#e0e0e0"),
        yaxis=dict(tickfont=dict(size=12, color="#222"), title_font=dict(size=13, color="#222"),
                   linecolor="#888", gridcolor="#e0e0e0"),
        paper_bgcolor="white", plot_bgcolor="#fafafa",
        legend=dict(font=dict(size=12, color="#111")),
    )
)

# ─── Constants ────────────────────────────────────────────────────────────────
_HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(_HERE, "data")

CAT_COLORS = {
    "BIOLOGICAL":   "#27ae60",
    "CLINICAL":     "#2980b9",
    "FUNCTIONAL":   "#e67e22",
    "SENSORY":      "#8e44ad",
    "UNCLASSIFIED": "#7f8c8d",
}

# ─── Project / data helpers ───────────────────────────────────────────────────

def load_index() -> list:
    """Return list of project dicts from data/index.json."""
    path = os.path.join(DATA_DIR, "index.json")
    if not os.path.exists(path):
        return []
    with open(path, encoding="utf-8") as f:
        return json.load(f).get("projects", [])


def load_meta(project_id: str) -> dict:
    """Return meta.json for a project, or sensible defaults."""
    path = os.path.join(DATA_DIR, project_id, "meta.json")
    defaults = {
        "project_id": project_id,
        "project_name": project_id,
        "group_a": "Group A",
        "group_b": "Group B",
        "researcher": "",
        "date": "",
        "notes": "",
        "institution": "",
        "lab": "",
    }
    if not os.path.exists(path):
        return defaults
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    return {**defaults, **data}


def project_paths(project_id: str) -> tuple:
    """Return (insight_path, groups_path) for a project."""
    base = os.path.join(DATA_DIR, project_id)
    return os.path.join(base, "insight.xlsx"), os.path.join(base, "groups.xlsx")


def detect_stat_columns(df: pd.DataFrame) -> tuple:
    """Auto-detect Log2FC, p-value, adj-p columns by pattern."""
    fc_col = next((c for c in df.columns if re.match(r'Log2 Fold Change', str(c), re.I)), None)
    pv_col = next((c for c in df.columns if re.match(r'P-value', str(c), re.I)
                   and 'adj' not in str(c).lower()), None)
    ap_col = next((c for c in df.columns if re.match(r'Adj\.?\s*P-value', str(c), re.I)), None)
    return fc_col, pv_col, ap_col


def groups_from_fc_col(fc_col) -> tuple:
    """'Log2 Fold Change: (Control) / (ADHD)' → ('Control', 'ADHD')"""
    if not fc_col:
        return "Group A", "Group B"
    m = re.search(r'\((.+?)\)\s*/\s*\((.+?)\)', fc_col)
    return (m.group(1), m.group(2)) if m else ("Group A", "Group B")


# ─── General helpers ──────────────────────────────────────────────────────────

def code_to_col(clinical_code: str) -> str:
    """'1A' → 'A1', '36B' → 'B36'"""
    m = re.match(r'^(\d+)([A-Za-z]+)$', str(clinical_code))
    return (m.group(2).upper() + m.group(1)) if m else str(clinical_code)


@st.cache_data(show_spinner="Loading group assignments…")
def load_group_map(path: str) -> dict:
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
    except Exception:
        return {}  # groups file missing or invalid — group-based features disabled
    # Try sheet 'ALL_data' first (ADHD project format), else use first sheet
    ws = wb['ALL_data'] if 'ALL_data' in wb.sheetnames else wb.active
    mapping = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = [v for v in row if v is not None]
        if len(row) >= 7:          # ADHD format: code at index 5, group at index 6
            code, group = row[5], row[6]
        elif len(row) >= 2:        # Generic format: code at index 0, group at index 1
            code, group = row[0], row[1]
        else:
            continue
        if code and group:
            mapping[code_to_col(str(code))] = str(group).strip().upper()
    return mapping


@st.cache_data(show_spinner="Parsing insights file…")
def load_insights(path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb['Sheet1']
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    rows    = list(ws.iter_rows(min_row=2, values_only=True))
    return pd.DataFrame(rows, columns=headers)


def detect_sample_cols(df: pd.DataFrame, group_map: dict = None) -> list:
    """Detect sample intensity columns. Uses group_map keys if available, else regex ^[A-Z]\\d+$."""
    if group_map:
        return [c for c in df.columns if str(c) in group_map]
    return [c for c in df.columns if isinstance(c, str) and re.match(r'^[A-Z]\d+$', c)]


def explode_tags(series: pd.Series) -> pd.Series:
    return (series.dropna()
                  .apply(lambda x: [t.strip() for t in str(x).split(';')
                                    if t.strip() and t.strip() != '-'])
                  .explode()
                  .reset_index(drop=True))


def apply_template(fig):
    fig.update_layout(**PLOT_TEMPLATE['layout'])
    return fig


# ─── Statistics helpers ───────────────────────────────────────────────────────

def _bh_correction(p_values: np.ndarray) -> np.ndarray:
    """Benjamini-Hochberg FDR correction (vectorised)."""
    n = len(p_values)
    if n == 0:
        return p_values.copy()
    order = np.argsort(p_values)
    adj = np.minimum(1.0, p_values[order] * n / (np.arange(n) + 1))
    # Enforce monotonicity (from right to left)
    np.minimum.accumulate(adj[::-1], out=adj[::-1])
    result = np.empty(n)
    result[order] = adj
    return result


@st.cache_data(show_spinner="Re-computing statistics from raw intensities…")
def compute_stats_live(
    data_json: str,
    adhd_cols: tuple,
    control_cols: tuple,
    method: str,
) -> tuple:
    """Recompute Log2FC, p-value, and BH-adjusted p from raw sample columns.

    FC convention: log2(median_group_b / median_group_a)
    Positive FC → elevated in group_b (numerator); negative → elevated in group_a (denominator).
    """
    df = pd.read_json(data_json, orient='split')
    adhd_cols  = list(adhd_cols)
    ctrl_cols  = list(control_cols)

    fc_arr  = np.full(len(df), np.nan)
    p_arr   = np.full(len(df), np.nan)

    for i, (_, row) in enumerate(df.iterrows()):
        a = pd.to_numeric(row[adhd_cols], errors='coerce').dropna().values
        c = pd.to_numeric(row[ctrl_cols], errors='coerce').dropna().values
        if len(a) < 2 or len(c) < 2:
            continue
        # p-value
        if method == "Mann-Whitney U":
            _, p = _scipy_stats.mannwhitneyu(a, c, alternative='two-sided')
        elif method == "Welch's t-test":
            _, p = _scipy_stats.ttest_ind(a, c, equal_var=False)
        else:  # Student's t-test
            _, p = _scipy_stats.ttest_ind(a, c, equal_var=True)
        p_arr[i] = p
        # Log2FC: median(group_b) / median(group_a) — same convention as precomputed
        med_a, med_c = np.median(a), np.median(c)
        if med_a > 0 and med_c > 0:
            fc_arr[i] = np.log2(med_c / med_a)

    valid = ~np.isnan(p_arr)
    adj_arr = np.full(len(df), np.nan)
    if valid.sum() > 0:
        adj_arr[valid] = _bh_correction(p_arr[valid])

    return fc_arr, p_arr, adj_arr


@st.cache_data(show_spinner="Running Shapiro-Wilk normality test…")
def run_normality_test(data_json: str, adhd_cols: tuple, control_cols: tuple) -> dict:
    """Run Shapiro-Wilk per compound; return summary stats."""
    df = pd.read_json(data_json, orient='split')
    all_cols = list(adhd_cols) + list(control_cols)
    normal_count = 0
    tested = 0
    p_values = []
    for _, row in df.iterrows():
        vals = pd.to_numeric(row[all_cols], errors='coerce').dropna().values
        if len(vals) >= 3:
            _, p = _scipy_stats.shapiro(vals)
            p_values.append(p)
            tested += 1
            if p > 0.05:
                normal_count += 1
    pct_normal = 100 * normal_count / tested if tested else 0
    return {
        "tested": tested,
        "normal": normal_count,
        "non_normal": tested - normal_count,
        "pct_normal": pct_normal,
        "median_p": float(np.median(p_values)) if p_values else np.nan,
    }


# ─── Download helpers ─────────────────────────────────────────────────────────

EXPORT_BASE = [
    'Original_Name', 'Formula_Final', 'MW_Final', 'HMDB_ID', 'KEGG_ID_Final',
    'Main_Categories', 'LM_Category', 'LM_Main_Class', 'LM_Sub_Class',
    'OC_Ratio', 'NOSC', 'KEGG_Pathways', 'Neuro_Trap', 'Clinical_Tags', 'Structural_Tags',
]


def to_csv(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False).encode('utf-8')


def to_excel(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        df.to_excel(w, index=False)
    return buf.getvalue()


def export_cols(df: pd.DataFrame, name_col: str, fc_col: str, pval_col: str) -> list:
    stat_cols = [fc_col, pval_col, '_direction']
    cols = [name_col] + stat_cols + [c for c in EXPORT_BASE if c in df.columns and c != name_col]
    return [c for c in cols if c in df.columns]


def build_metaboanalyst(df: pd.DataFrame, sample_cols: list, grp_norm: dict):
    """Build MetaboAnalyst peak table and metadata DataFrames."""
    def best_id(row):
        pubchem = str(row.get('PubChem_CID', '') or '').strip()
        hmdb    = str(row.get('HMDB_ID',     '') or '').strip()
        kegg    = str(row.get('KEGG_ID_Final','') or '').strip()
        if pubchem and pubchem not in ('nan', '-', ''):
            try:
                return str(int(float(pubchem)))
            except (ValueError, OverflowError):
                pass
        if hmdb.startswith('HMDB') and len(hmdb) > 4:
            return hmdb
        if kegg and kegg not in ('-', 'nan', ''):
            return kegg
        return None

    df2 = df.copy()
    df2['_ma_id'] = df2.apply(best_id, axis=1)
    included   = df2[df2['_ma_id'].notna()].copy()
    n_excluded = len(df2) - len(included)

    peak_rows = []
    for _, row in included.iterrows():
        r = {'PubChem_CID': row['_ma_id']}
        for s in sample_cols:
            val = pd.to_numeric(row.get(s, np.nan), errors='coerce')
            r[s] = 0 if pd.isna(val) else val
        peak_rows.append(r)
    peak_table = pd.DataFrame(peak_rows)

    meta = pd.DataFrame([
        {'Sample': s, 'Group': grp_norm.get(s, 'Unknown')}
        for s in sample_cols
    ])
    return peak_table, meta, len(included), n_excluded


# ─── Main app ─────────────────────────────────────────────────────────────────

def main():
    # ── Resolve project ───────────────────────────────────────────────────────
    params      = st.query_params
    url_project = params.get("project", None)
    index       = load_index()
    proj_ids    = [p["project_id"] for p in index]

    with st.sidebar:
        st.header("📂 Project")
        if proj_ids:
            if url_project and url_project in proj_ids:
                default_idx = proj_ids.index(url_project)
            else:
                default_idx = 0
            selected_id = st.selectbox(
                "Select project",
                options=proj_ids,
                index=default_idx,
                format_func=lambda pid: next(
                    (p["project_name"] for p in index if p["project_id"] == pid), pid
                ),
            )
            # Sync URL param
            if st.query_params.get("project") != selected_id:
                st.query_params["project"] = selected_id
        else:
            selected_id = url_project or "adhd_2026"
            st.caption(f"Project: {selected_id}")

        meta = load_meta(selected_id)
        with st.expander("ℹ️ Project info"):
            st.caption(f"**{meta['project_name']}**")
            if meta['researcher']: st.caption(f"Researcher: {meta['researcher']}")
            if meta['date']:       st.caption(f"Date: {meta['date']}")
            if meta['notes']:      st.caption(meta['notes'])

        st.divider()
        _comparison_container = st.container()   # filled later after data is loaded
        st.divider()
        st.header("🧪 Statistical Method")
        STAT_OPTIONS = [
            "Precomputed (Mann-Whitney U)",
            "Mann-Whitney U",
            "Welch's t-test",
            "Student's t-test",
        ]
        STAT_DESC = {
            "Precomputed (Mann-Whitney U)":
                "Non-parametric rank-sum test. No normality assumption. "
                "Standard choice for untargeted metabolomics (Wilcoxon Mann-Whitney, two-sided).",
            "Mann-Whitney U":
                "Recomputed from raw intensities. Same non-parametric test; "
                "use to verify or update after data changes.",
            "Welch's t-test":
                "Parametric. Allows unequal variances. Appropriate when data are "
                "approximately normally distributed but group SDs differ.",
            "Student's t-test":
                "Parametric. Assumes equal variances AND normality. "
                "Generally not recommended for metabolomics without prior normality confirmation.",
        }
        stat_method = st.selectbox("Test", STAT_OPTIONS, index=0)
        with st.expander("ℹ️ About this test"):
            st.caption(STAT_DESC[stat_method])

        run_norm_btn = st.button("🔬 Run Normality Test (Shapiro-Wilk)",
                                 help="Tests each compound across all samples. Takes ~5 s.")

        st.divider()
        st.header("🔬 Statistical Filters")
        fc_thresh   = st.slider("|Log₂ FC| threshold", 0.0, 4.0, 1.0, 0.25,
                                help="Minimum absolute fold-change for significance")
        pval_thresh = st.selectbox("P-value threshold",
                                   [1.0, 0.5, 0.1, 0.05, 0.01, 0.001],
                                   index=3,
                                   format_func=lambda v: "Any (no filter)" if v == 1.0 else str(v))
        show_adj    = st.checkbox("Use adjusted p-value (FDR)", value=False)

        st.divider()
        _footer_inst = meta.get('institution', '') or 'MetaFlow'
        st.caption(f"MetaFlow · {_footer_inst} · {meta.get('date', '2026')[:4]}")

    # ── Load data ────────────────────────────────────────────────────────────
    insights_path, groups_path = project_paths(selected_id)

    group_map   = load_group_map(groups_path)
    df          = load_insights(insights_path)
    sample_cols = detect_sample_cols(df, group_map)

    # ── Dynamic group names ───────────────────────────────────────────────────
    fc_col, pval_col, adjp_col = detect_stat_columns(df)
    fc_col   = fc_col   or 'Log2 Fold Change: (Control) / (ADHD)'
    pval_col = pval_col or 'P-value: (Control) / (ADHD)'
    adjp_col = adjp_col or 'Adj. P-value: (Control) / (ADHD)'

    # Group names from meta.json (authoritative) or inferred from FC column header
    inferred_num, inferred_den = groups_from_fc_col(fc_col)
    GROUP_A = meta.get("group_a") or inferred_den   # denominator = "elevated in GROUP_A" when FC < 0
    GROUP_B = meta.get("group_b") or inferred_num   # numerator

    # ── Build grp_norm preserving ALL group names ─────────────────────────────
    # Canonical names from any group_* field in meta.json (covers group_a, group_b, group_c …)
    meta_canonical = {meta[k].upper(): meta[k]
                      for k in meta if k.startswith('group_') and meta.get(k)}
    grp_norm = {k: meta_canonical.get(v, v) for k, v in group_map.items()}
    all_groups = sorted(set(grp_norm.values()), key=str.casefold)

    # ── Comparison selector (fills the sidebar container reserved above) ────────
    _PALETTE = ["#c0392b", "#1a6ea8", "#27ae60", "#e67e22", "#8e44ad", "#16a085", "#f39c12"]
    _other   = [g for g in all_groups if g not in (GROUP_A, GROUP_B)]
    COLORS   = {"neutral": "#7f8c8d"}
    for _i, _g in enumerate([GROUP_A, GROUP_B] + _other):
        COLORS[_g] = _PALETTE[_i % len(_PALETTE)]

    if len(all_groups) > 2:
        _default_a = [g for g in [GROUP_A] if g in all_groups] or [all_groups[0]]
        _default_b = [g for g in [GROUP_B] if g in all_groups] or [all_groups[min(1, len(all_groups)-1)]]
        with _comparison_container:
            st.subheader("🔀 Comparison Groups")
            selected_a = st.multiselect("Group A (statistics)",
                                        all_groups, default=_default_a, key='cmp_a',
                                        help="Groups used as Group A in FC and significance tests")
            selected_b = st.multiselect("Group B (statistics)",
                                        all_groups, default=_default_b, key='cmp_b',
                                        help="Groups used as Group B in FC and significance tests")
            # remaining groups offered as display-only
            _not_in_ab = [g for g in all_groups if g not in selected_a and g not in selected_b]
            display_extra = st.multiselect("Additional groups (display only)",
                                           _not_in_ab, default=_not_in_ab, key='cmp_extra',
                                           help="Shown in charts alongside A and B — not included in statistics")
            _overlap = set(selected_a) & set(selected_b)
            if _overlap:
                st.warning(f"Overlap: {', '.join(_overlap)} is in both A and B.")
            if not selected_a or not selected_b:
                st.info("Select at least one group per side.")
                selected_a = selected_a or _default_a
                selected_b = selected_b or _default_b
        # Labels for stat-facing variables
        GROUP_A = " + ".join(selected_a)
        GROUP_B = " + ".join(selected_b)
        COLORS[GROUP_A] = COLORS.get(selected_a[0], _PALETTE[0])
        COLORS[GROUP_B] = COLORS.get(selected_b[0], _PALETTE[1])
        adhd_cols    = [c for c in sample_cols if grp_norm.get(c) in selected_a]
        control_cols = [c for c in sample_cols if grp_norm.get(c) in selected_b]
        extra_cols   = [c for c in sample_cols if grp_norm.get(c) in display_extra]
    else:
        selected_a   = [GROUP_A]
        selected_b   = [GROUP_B]
        display_extra = []
        extra_cols   = []

    DIR_A = f"↑ {GROUP_A}"
    DIR_B = f"↑ {GROUP_B}"

    # ── Page header ──────────────────────────────────────────────────────────
    proj_title = meta.get("project_name", "Metabolomics Dashboard")
    _institution = meta.get('institution', '')
    _lab         = meta.get('lab', '')
    if len(all_groups) > 2:
        _groups_str = " · ".join(all_groups)
    else:
        _groups_str = f"{GROUP_A} vs {GROUP_B}"
    _subtitle_parts = [p for p in [_groups_str, _institution, _lab] if p]
    proj_subtitle = " · ".join(_subtitle_parts)
    st.markdown(f"""
    <div class="main-header">
        <h1>🧬 {proj_title}</h1>
        <p>{proj_subtitle}</p>
    </div>""", unsafe_allow_html=True)

    # ── Sample columns per comparison (2-group case; multi-group set above) ─────
    if len(all_groups) <= 2:
        adhd_cols    = [c for c in sample_cols if grp_norm.get(c) in selected_a]
        control_cols = [c for c in sample_cols if grp_norm.get(c) in selected_b]

    name_col = 'Name' if 'Name' in df.columns else 'Original_Name'

    for col in [fc_col, pval_col, adjp_col, 'OC_Ratio', 'NOSC']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # ── Recompute stats if non-default method selected ────────────────────────
    if stat_method != "Precomputed (Mann-Whitney U)" and adhd_cols and control_cols:
        _data_json = df.to_json(orient='split')
        fc_arr, p_arr, adj_arr = compute_stats_live(
            _data_json, tuple(adhd_cols), tuple(control_cols), stat_method
        )
        df = df.copy()
        df[fc_col]   = fc_arr
        df[pval_col] = p_arr
        df[adjp_col] = adj_arr

    # ── Normality test ────────────────────────────────────────────────────────
    if run_norm_btn and adhd_cols and control_cols:
        _data_json = df.to_json(orient='split')
        norm = run_normality_test(_data_json, tuple(adhd_cols), tuple(control_cols))
        pct = norm['pct_normal']
        rec = "Mann-Whitney U" if pct < 70 else "Welch's t-test or Mann-Whitney U"
        st.info(
            f"**Shapiro-Wilk Normality Test** — {norm['tested']} compounds tested\n\n"
            f"- Normal (p > 0.05): **{norm['normal']}** ({pct:.1f}%)\n"
            f"- Non-normal: **{norm['non_normal']}** ({100-pct:.1f}%)\n"
            f"- Median p-value: {norm['median_p']:.4f}\n\n"
            f"**Recommendation:** {rec}"
        )

    p_use = adjp_col if show_adj else pval_col

    # pre-compute significance direction once
    df['_direction'] = df.apply(
        lambda r: (DIR_A if r.get(fc_col, np.nan) < -fc_thresh
                   else (DIR_B if r.get(fc_col, np.nan) > fc_thresh else 'n/s'))
        if (pd.notna(r.get(fc_col)) and pd.notna(r.get(p_use))
            and abs(r.get(fc_col, 0)) >= fc_thresh
            and r.get(p_use, 1) < pval_thresh)
        else 'n/s', axis=1)

    # ── TABS ─────────────────────────────────────────────────────────────────
    tabs = st.tabs([
        "📊 Overview",
        "🧪 Lipids",
        "⚡ Oxidation",
        "🗺 KEGG Pathways",
        "🧠 Neurochemistry",
        f"📈 {GROUP_A} vs {GROUP_B}",
        "🔍 Compound Explorer",
        "📥 Downloads & Exports",
    ])

    # ═════════════════════════════════════════════════════════════════════════
    # TAB 1 — OVERVIEW
    # ═════════════════════════════════════════════════════════════════════════
    with tabs[0]:
        n_sig_a = int((df['_direction'] == DIR_A).sum())
        n_sig_b = int((df['_direction'] == DIR_B).sum())

        metrics = [
            (len(df),                                          "Total Compounds"),
            (len(adhd_cols),                                   f"{GROUP_A} Samples"),
            (len(control_cols),                                f"{GROUP_B} Samples"),
            (int(df['LM_Category'].notna().sum()),             "Classified Lipids"),
            (int((df['KEGG_Pathways'].notna() & (df['KEGG_Pathways'] != '-')).sum()), "In KEGG Pathways"),
            (int((df['Neuro_Trap'].notna() & (df['Neuro_Trap'] != '-')).sum()),        "Neuro-Active"),
        ]
        cols = st.columns(6)
        for col, (val, label) in zip(cols, metrics):
            col.markdown(f"""<div class="metric-card">
                <div class="value">{val:,}</div>
                <div class="label">{label}</div></div>""", unsafe_allow_html=True)

        st.markdown("<div class='section-header'>Functional Classification of All Compounds</div>",
                    unsafe_allow_html=True)
        cl, cr = st.columns(2)

        # Pie: main categories
        cat_ctr = Counter()
        for v in df['Main_Categories'].dropna():
            for c in str(v).split(';'):
                c = c.strip()
                if c: cat_ctr[c] += 1
        cat_df = pd.DataFrame(cat_ctr.items(), columns=['Category','Count'])
        fig_pie = px.pie(cat_df, names='Category', values='Count',
                         color='Category', color_discrete_map=CAT_COLORS,
                         title='Compound Categories (multi-label)', hole=0.42)
        fig_pie.update_traces(textinfo='percent+label', textfont_size=13,
                              pull=[0.04]*len(cat_df))
        fig_pie.update_layout(showlegend=False, margin=dict(t=50,b=10))
        apply_template(fig_pie)
        cl.plotly_chart(fig_pie, use_container_width=True)

        # Bar: structural tags
        stag = explode_tags(df['Structural_Tags'])
        stag_df = stag.value_counts().head(15).reset_index()
        stag_df.columns = ['Tag','Count']
        stag_df['Tag'] = stag_df['Tag'].str.split(':').str[-1].str.replace('_',' ')
        fig_stag = px.bar(stag_df, x='Count', y='Tag', orientation='h',
                          color='Count', color_continuous_scale='Blues',
                          title='Top 15 Structural / Biological Tags')
        fig_stag.update_layout(yaxis={'categoryorder':'total ascending'},
                                coloraxis_showscale=False, margin=dict(t=50,b=10))
        apply_template(fig_stag)
        cr.plotly_chart(fig_stag, use_container_width=True)

        # Significance overview
        st.markdown("<div class='section-header'>Significance Summary (current filter thresholds)</div>",
                    unsafe_allow_html=True)
        sc1, sc2, sc3 = st.columns(3)
        sc1.markdown(f"""<div class="metric-card">
            <div class="value" style="color:#c0392b">{n_sig_a}</div>
            <div class="label">↑ Elevated in {GROUP_A}</div></div>""", unsafe_allow_html=True)
        sc2.markdown(f"""<div class="metric-card">
            <div class="value" style="color:#1a6ea8">{n_sig_b}</div>
            <div class="label">↑ Elevated in {GROUP_B}</div></div>""", unsafe_allow_html=True)
        sc3.markdown(f"""<div class="metric-card">
            <div class="value" style="color:#555">{len(df)-n_sig_a-n_sig_b:,}</div>
            <div class="label">Not Significant</div></div>""", unsafe_allow_html=True)

        # Sample map
        st.markdown("<div class='section-header'>Sample Group Map</div>", unsafe_allow_html=True)
        smap = pd.DataFrame([
            {'Sample': s, 'Group': grp_norm.get(s, 'Unknown')}
            for s in sample_cols
        ])
        fig_map = px.strip(smap, x='Sample', color='Group',
                           color_discrete_map=COLORS,
                           category_orders={'Group': [GROUP_A, GROUP_B]},
                           title=f'{len(sample_cols)} samples — {len(adhd_cols)} {GROUP_A} · {len(control_cols)} {GROUP_B}')
        fig_map.update_traces(marker_size=14, jitter=0)
        fig_map.update_layout(xaxis_tickangle=45, margin=dict(t=50,b=90),
                               xaxis_tickfont_size=11)
        apply_template(fig_map)
        st.plotly_chart(fig_map, use_container_width=True)

        with st.expander(f"📋 Full sample list by group"):
            ea, ec = st.columns(2)
            ea.markdown(f"**{GROUP_A} samples**")
            ea.write(sorted(adhd_cols))
            ec.markdown(f"**{GROUP_B} samples**")
            ec.write(sorted(control_cols))

    # ═════════════════════════════════════════════════════════════════════════
    # TAB 2 — LIPIDS
    # ═════════════════════════════════════════════════════════════════════════
    with tabs[1]:
        lipid_df = df[df['LM_Category'].notna()].copy()
        st.markdown(f"### {len(lipid_df):,} classified lipids (LIPID MAPS annotation)")

        cl, cr = st.columns(2)

        sb_df = lipid_df[['LM_Category','LM_Main_Class','LM_Sub_Class']].fillna('Unknown')
        sb_df = sb_df[sb_df['LM_Category'] != 'Unknown'].copy()
        sb_df['n'] = 1
        fig_sun = px.sunburst(sb_df, path=['LM_Category','LM_Main_Class','LM_Sub_Class'],
                              values='n', title='Lipid Class Hierarchy — Sunburst',
                              color_discrete_sequence=px.colors.qualitative.Set2)
        fig_sun.update_layout(margin=dict(t=50,b=10),
                               title_font_size=15)
        cl.plotly_chart(fig_sun, use_container_width=True)

        lm_mc = lipid_df['LM_Main_Class'].value_counts().head(18).reset_index()
        lm_mc.columns = ['Class','Count']
        lm_mc['Class'] = lm_mc['Class'].str.replace(r'\s*\[.*?\]', '', regex=True)
        fig_lm = px.bar(lm_mc, x='Count', y='Class', orientation='h',
                        color='Count', color_continuous_scale='Teal',
                        title='Lipid Main Classes (top 18)')
        fig_lm.update_layout(yaxis={'categoryorder':'total ascending'},
                              coloraxis_showscale=False, margin=dict(t=50,b=10))
        apply_template(fig_lm)
        cr.plotly_chart(fig_lm, use_container_width=True)

        st.markdown("<div class='section-header'>Lipid Sub-Classes — Expand for Group Comparison</div>",
                    unsafe_allow_html=True)
        lm_sub_top = lipid_df['LM_Sub_Class'].value_counts()
        for _lm_idx, (sub_class, count) in enumerate(lm_sub_top.items()):
            if sub_class in ('-','Unknown') or not sub_class: continue
            sub_df = lipid_df[lipid_df['LM_Sub_Class'] == sub_class]
            short  = re.sub(r'\s*\[.*?\]','', sub_class)
            n_a_up = int((sub_df['_direction'] == DIR_A).sum())
            n_b_up = int((sub_df['_direction'] == DIR_B).sum())
            n_ns   = count - n_a_up - n_b_up
            med_fc = sub_df[fc_col].median() if fc_col in sub_df.columns else np.nan
            with st.expander(f"**{short}** — {count} compounds"):
                g1, g2 = st.columns(2)
                with g1:
                    st.markdown("**Directional compound counts:**")
                    st.markdown(f"<span class='badge-adhd'>{DIR_A}: {n_a_up}</span>&nbsp;&nbsp;"
                                f"<span class='badge-ctrl'>{DIR_B}: {n_b_up}</span>&nbsp;&nbsp;"
                                f"<span style='background:#888;color:white;padding:3px 8px;border-radius:12px;"
                                f"font-size:0.85rem;font-weight:600'>n/s: {n_ns}</span>",
                                unsafe_allow_html=True)
                    if np.isfinite(med_fc):
                        direction = DIR_B if med_fc > 0 else DIR_A
                        st.metric(f"Median Log₂FC ({GROUP_B}/{GROUP_A})",
                                  f"{med_fc:+.2f}",
                                  help=f"Positive = higher in {GROUP_B}. {direction}")
                    # small mirror bar
                    if n_a_up > 0 or n_b_up > 0:
                        mini = pd.DataFrame({
                            'Group': [DIR_A, DIR_B],
                            'Count': [-n_a_up, n_b_up],
                            'color': [COLORS[GROUP_A], COLORS[GROUP_B]]
                        })
                        fig_mini = px.bar(mini, x='Count', y='Group', orientation='h',
                                          color='Group',
                                          color_discrete_map={DIR_A: COLORS[GROUP_A],
                                                              DIR_B: COLORS[GROUP_B]},
                                          title='Significant compounds')
                        fig_mini.update_layout(showlegend=False, height=120,
                                               margin=dict(t=30,b=5,l=5,r=5),
                                               xaxis=dict(tickformat='d', zeroline=True,
                                                          zerolinecolor='#333', zerolinewidth=2),
                                               font=dict(size=12,color='#111'),
                                               paper_bgcolor='white', plot_bgcolor='#fafafa')
                        st.plotly_chart(fig_mini, use_container_width=True,
                                        key=f"lipid_mini_{_lm_idx}")
                with g2:
                    names = sub_df[name_col].dropna().unique()
                    st.markdown(f"**{len(names)} compounds:**")
                    st.write(list(names[:30]))

    # ═════════════════════════════════════════════════════════════════════════
    # TAB 3 — OXIDATION
    # ═════════════════════════════════════════════════════════════════════════
    with tabs[2]:
        st.markdown("### Oxidative Stress Markers")

        oc_df = df[df['OC_Ratio'].notna()].copy()
        oc_df['Oxidation_Level'] = pd.cut(oc_df['OC_Ratio'],
            bins=[-np.inf, 0.2, 0.4, 0.6, 0.8, np.inf],
            labels=['Very Low','Low','Medium','High','Very High'])

        cl, cr = st.columns(2)

        fig_oc = px.histogram(oc_df, x='OC_Ratio', nbins=60,
                              color_discrete_sequence=['#c0392b'],
                              title='O/C Ratio Distribution — Oxidation Index',
                              labels={'OC_Ratio':'O/C Ratio', 'count':'# Compounds'})
        fig_oc.add_vline(x=0.6, line_dash='dash', line_color='#1a1a8c',
                         annotation_text='  High oxidation threshold (0.6)',
                         annotation_font_size=12)
        fig_oc.update_layout(margin=dict(t=50,b=10))
        apply_template(fig_oc)
        cl.plotly_chart(fig_oc, use_container_width=True)

        ox_c = oc_df['Oxidation_Level'].value_counts().reset_index()
        ox_c.columns = ['Level','Count']
        fig_ox_pie = px.pie(ox_c, names='Level', values='Count',
                            color='Level',
                            color_discrete_sequence=['#2ecc71','#f1c40f','#e67e22','#e74c3c','#8e44ad'],
                            title='Compounds by Oxidation Level', hole=0.4)
        fig_ox_pie.update_traces(
            textinfo='percent+label',
            textfont_size=11,
            textposition='outside',
            outsidetextfont=dict(size=11),
        )
        fig_ox_pie.update_layout(
            margin=dict(t=50, b=80, l=60, r=60),
            showlegend=True,
            legend=dict(orientation='h', y=-0.18, x=0.5, xanchor='center', font_size=11),
        )
        cr.plotly_chart(fig_ox_pie, use_container_width=True)

        st.markdown("<div class='section-header'>Clinical Oxidative Stress Categories</div>",
                    unsafe_allow_html=True)
        ox_kw = ['Oxylipin','Epoxide','Oxidative','Reactive','Peroxide']
        ox_dict = Counter()
        for v in df['Clinical_Tags'].dropna():
            for tag in str(v).split(';'):
                tag = tag.strip()
                if any(k in tag for k in ox_kw):
                    ox_dict[tag.split(':')[-1].replace('_',' ')] += 1

        if ox_dict:
            ox_df2 = pd.DataFrame(ox_dict.most_common(12), columns=['Type','Count'])
            fig_oxt = px.bar(ox_df2, x='Type', y='Count',
                             color='Count', color_continuous_scale='Reds',
                             title='Oxidative Stress Compound Types')
            fig_oxt.update_layout(coloraxis_showscale=False,
                                   margin=dict(t=50,b=90), xaxis_tickangle=30)
            apply_template(fig_oxt)
            st.plotly_chart(fig_oxt, use_container_width=True)

        st.markdown(f"<div class='section-header'>High-Oxidation Compounds (O/C > 0.6) — {GROUP_A} vs {GROUP_B}</div>",
                    unsafe_allow_html=True)
        high_ox = oc_df[oc_df['OC_Ratio'] > 0.6].copy()
        # Direction counts for high-ox pool
        n_hox_a  = int((high_ox['_direction'] == DIR_A).sum())
        n_hox_b  = int((high_ox['_direction'] == DIR_B).sum())
        n_hox_ns = len(high_ox) - n_hox_a - n_hox_b

        hox_c1, hox_c2, hox_c3 = st.columns(3)
        hox_c1.markdown(f"""<div class="metric-card">
            <div class="value" style="color:#c0392b">{n_hox_a}</div>
            <div class="label">{DIR_A} (sig.)</div></div>""", unsafe_allow_html=True)
        hox_c2.markdown(f"""<div class="metric-card">
            <div class="value" style="color:#1a6ea8">{n_hox_b}</div>
            <div class="label">{DIR_B} (sig.)</div></div>""", unsafe_allow_html=True)
        hox_c3.markdown(f"""<div class="metric-card">
            <div class="value" style="color:#555">{n_hox_ns}</div>
            <div class="label">Not significant</div></div>""", unsafe_allow_html=True)

        # Log2FC bar chart for top high-OC compounds (only those with FC data)
        hox_fc = high_ox[high_ox[fc_col].notna() & high_ox[pval_col].notna()].copy()
        hox_fc = hox_fc.drop_duplicates(subset=[name_col]).nlargest(30, 'OC_Ratio')
        hox_fc['label_color'] = hox_fc['_direction'].map(
            {DIR_A: COLORS[GROUP_A], DIR_B: COLORS[GROUP_B], 'n/s': '#aaa'})
        hox_fc_sorted = hox_fc.sort_values(fc_col)
        fig_hox = go.Figure(go.Bar(
            x=hox_fc_sorted[fc_col],
            y=hox_fc_sorted[name_col],
            orientation='h',
            marker_color=hox_fc_sorted['label_color'].tolist(),
            hovertemplate='<b>%{y}</b><br>Log₂FC: %{x:.2f}<extra></extra>',
        ))
        fig_hox.add_vline(x=0, line_color='#333', line_width=1.5)
        fig_hox.add_vline(x= fc_thresh, line_dash='dash', line_color='#888', line_width=1)
        fig_hox.add_vline(x=-fc_thresh, line_dash='dash', line_color='#888', line_width=1)
        fig_hox.update_layout(
            title=dict(text=f'Top 30 High-Oxidation Compounds (O/C > 0.6) — Log₂ Fold Change<br>'
                            f'<sup style="color:{COLORS[GROUP_A]}">■ {DIR_A}</sup>&nbsp;&nbsp;'
                            f'<sup style="color:{COLORS[GROUP_B]}">■ {DIR_B}</sup>&nbsp;&nbsp;'
                            '<sup style="color:#aaa">■ n/s</sup>', font_size=14),
            xaxis=dict(title=f'Log₂ FC ({GROUP_B} / {GROUP_A})', tickfont_size=12,
                       zeroline=True, zerolinecolor='#333'),
            yaxis=dict(tickfont_size=11),
            height=max(400, len(hox_fc_sorted) * 20),
            margin=dict(t=80, b=30, l=10, r=20),
            paper_bgcolor='white', plot_bgcolor='#fafafa',
            font=dict(size=12, color='#111'),
        )
        st.plotly_chart(fig_hox, use_container_width=True)

        with st.expander(f"📋 All {len(high_ox)} highly-oxidised compounds (O/C > 0.6)"):
            show = [name_col,'OC_Ratio','NOSC', fc_col, pval_col, '_direction','Main_Categories','Clinical_Tags']
            show = [c for c in show if c in high_ox.columns]
            st.dataframe(high_ox[show].rename(columns={fc_col:'Log2FC', pval_col:'p-value',
                                                        '_direction':'Direction'})
                         .reset_index(drop=True), use_container_width=True)

    # ═════════════════════════════════════════════════════════════════════════
    # TAB 4 — KEGG PATHWAYS
    # ═════════════════════════════════════════════════════════════════════════
    with tabs[3]:
        kegg_df = df[df['KEGG_Pathways'].notna() & (df['KEGG_Pathways'] != '-')].copy()

        # ── build per-pathway counts ──────────────────────────────────────
        path_all = Counter()
        path_a   = Counter()   # ↑ GROUP_A (higher in research group)
        path_b   = Counter()   # ↑ GROUP_B (lower in research group)
        for _, row in kegg_df.iterrows():
            paths = [p.strip() for p in str(row['KEGG_Pathways']).split(';') if p.strip()]
            direction = row.get('_direction', '')
            for p in paths:
                path_all[p] += 1
                if direction == DIR_A: path_a[p] += 1
                if direction == DIR_B: path_b[p] += 1

        all_pathway_names = [p for p, _ in path_all.most_common()]
        n_sig_kegg = int((kegg_df['_direction'] != 'n/s').sum())

        pv_label = "Any" if pval_thresh >= 1.0 else f"p<{pval_thresh}"
        fc_label = f"|FC|≥{fc_thresh}" if fc_thresh > 0 else "any FC"

        st.markdown(
            f"### {len(kegg_df):,} compounds mapped to KEGG pathways "
            f"— **{n_sig_kegg}** significant ({fc_label}, {pv_label})"
        )

        # ── pathway selector ──────────────────────────────────────────────
        selected_paths = st.multiselect(
            "Pathways to display (add or remove freely):",
            options=all_pathway_names,
            default=all_pathway_names[:25],
            key='kegg_multiselect',
        )

        if not selected_paths:
            st.info("Select at least one pathway to display the chart.")
        else:
            pw_df = pd.DataFrame({
                'Pathway': selected_paths,
                'A_n':     [path_a[p] for p in selected_paths],   # ↑ GROUP_A (right)
                'B_n':     [path_b[p] for p in selected_paths],   # ↑ GROUP_B (left)
                'Total':   [path_all[p] for p in selected_paths],
            })
            pw_df['Sig_total'] = pw_df['A_n'] + pw_df['B_n']
            pw_df = pw_df.sort_values('Sig_total')   # most significant at top

            bar_h = max(28, min(48, 600 // max(len(pw_df), 1)))
            fig_h = max(420, len(pw_df) * bar_h + 160)

            fig_mirror = go.Figure()

            # LEFT bars = ↑ GROUP_B = lower in research group (negative x trick)
            fig_mirror.add_trace(go.Bar(
                name=f'↓ {GROUP_A}  (higher in {GROUP_B})',
                x=[-n for n in pw_df['B_n']],
                y=pw_df['Pathway'],
                orientation='h',
                marker=dict(color=COLORS[GROUP_B], line=dict(color='white', width=0.8)),
                customdata=pw_df['B_n'].values,
                hovertemplate=(f'<b>%{{y}}</b><br>'
                               f'Lower in {GROUP_A}: <b>%{{customdata}}</b><extra></extra>'),
                text=[str(n) if n > 0 else '' for n in pw_df['B_n']],
                textposition='inside',
                textfont=dict(size=12, color='white'),
                insidetextanchor='middle',
            ))

            # RIGHT bars = ↑ GROUP_A = higher in research group
            fig_mirror.add_trace(go.Bar(
                name=f'↑ {GROUP_A}  (lower in {GROUP_B})',
                x=pw_df['A_n'],
                y=pw_df['Pathway'],
                orientation='h',
                marker=dict(color=COLORS[GROUP_A], line=dict(color='white', width=0.8)),
                customdata=pw_df['A_n'].values,
                hovertemplate=(f'<b>%{{y}}</b><br>'
                               f'Higher in {GROUP_A}: <b>%{{customdata}}</b><extra></extra>'),
                text=[str(n) if n > 0 else '' for n in pw_df['A_n']],
                textposition='inside',
                textfont=dict(size=12, color='white'),
                insidetextanchor='middle',
            ))

            # n=total annotation on right margin
            x_max = max(pw_df['A_n'].max(), pw_df['B_n'].max(), 1)
            x_pad = x_max * 0.25
            for _, row_pw in pw_df.iterrows():
                fig_mirror.add_annotation(
                    x=x_max + x_pad * 0.6,
                    y=row_pw['Pathway'],
                    text=f"n={row_pw['Total']}",
                    showarrow=False,
                    font=dict(size=10, color='#666'),
                    xanchor='left',
                )

            fig_mirror.update_layout(
                barmode='relative',
                title=dict(
                    text=(f'<b>{GROUP_A} relative to {GROUP_B}</b>'
                          f'<br><span style="font-size:12px;color:#555">'
                          f'← Lower in {GROUP_A} &nbsp;|&nbsp; Higher in {GROUP_A} →'
                          f'&nbsp;&nbsp;({fc_label}, {pv_label})'
                          f'&nbsp;|&nbsp; n = total mapped</span>'),
                    font=dict(size=16, color='#111'),
                    x=0.5, xanchor='center',
                ),
                xaxis=dict(
                    title=dict(text='Number of significant metabolites', font=dict(size=13)),
                    tickfont=dict(size=12),
                    zeroline=True, zerolinecolor='#222', zerolinewidth=2,
                    gridcolor='#e8e8e8',
                    range=[-(x_max + x_pad), x_max + x_pad * 1.8],
                    tickformat='d',
                    labelalias={str(-v): str(v) for v in range(1, x_max + 2)},
                ),
                yaxis=dict(tickfont=dict(size=12, color='#111'), automargin=True),
                legend=dict(
                    orientation='h', x=0.5, y=-0.08, xanchor='center',
                    font=dict(size=13), bgcolor='rgba(0,0,0,0)',
                    itemsizing='constant',
                ),
                height=fig_h,
                margin=dict(t=90, b=70, l=20, r=90),
                paper_bgcolor='white', plot_bgcolor='white',
                font=dict(family='Arial, sans-serif', size=13, color='#111'),
                bargap=0.35,
            )

            for i, pathway in enumerate(pw_df['Pathway']):
                if i % 2 == 0:
                    fig_mirror.add_hrect(
                        y0=i - 0.5, y1=i + 0.5,
                        fillcolor='rgba(0,0,0,0.025)', line_width=0,
                    )

            st.plotly_chart(fig_mirror, use_container_width=True, key='kegg_mirror')

            # ── pathway detail expanders ──────────────────────────────────
            st.markdown("<div class='section-header'>Pathway Detail — Expand to see compound list</div>",
                        unsafe_allow_html=True)
            for path in sorted(selected_paths, key=lambda p: -path_all[p]):
                path_cpds = kegg_df[kegg_df['KEGG_Pathways'].str.contains(re.escape(path), na=False)]
                a_up      = path_cpds[path_cpds['_direction'] == DIR_A]
                b_up      = path_cpds[path_cpds['_direction'] == DIR_B]
                label = (f"**{path}** — {path_all[path]} total "
                         f"| {DIR_A}: {path_a[path]} | {DIR_B}: {path_b[path]}")
                with st.expander(label):
                    ec1, ec2 = st.columns(2)
                    with ec1:
                        st.markdown(f"<span class='badge-adhd'>{DIR_A} ({len(a_up)})</span>",
                                    unsafe_allow_html=True)
                        st.write(list(a_up[name_col].dropna().unique()))
                    with ec2:
                        st.markdown(f"<span class='badge-ctrl'>{DIR_B} ({len(b_up)})</span>",
                                    unsafe_allow_html=True)
                        st.write(list(b_up[name_col].dropna().unique()))
                    st.markdown("**All compounds in this pathway:**")
                    st.write(list(path_cpds[name_col].dropna().unique()))

    # ═════════════════════════════════════════════════════════════════════════
    # TAB 5 — NEUROCHEMISTRY
    # ═════════════════════════════════════════════════════════════════════════
    with tabs[4]:
        neuro_df = df[df['Neuro_Trap'].notna() & (df['Neuro_Trap'] != '-')].copy()
        st.markdown(f"### {len(neuro_df):,} neuro-active compounds detected")

        neuro_ctr = Counter()
        for v in neuro_df['Neuro_Trap']:
            for t in str(v).split(';'):
                t = t.strip()
                if t: neuro_ctr[t] += 1

        neuro_top = pd.DataFrame(neuro_ctr.most_common(20), columns=['Axis','Count'])
        cl, cr = st.columns([3,2])

        fig_neuro = px.bar(neuro_top, x='Count', y='Axis', orientation='h',
                           color='Count', color_continuous_scale='Purples',
                           title='Neuro-active Compound Axes')
        fig_neuro.update_layout(yaxis={'categoryorder':'total ascending'},
                                 coloraxis_showscale=False, margin=dict(t=50,b=10))
        apply_template(fig_neuro)
        cl.plotly_chart(fig_neuro, use_container_width=True)

        major_axes = [a for a in neuro_ctr if 'axis' in a.lower() or '[Pathway' in a][:8]
        if major_axes:
            rv = [neuro_ctr[a] for a in major_axes]
            fig_radar = go.Figure(go.Scatterpolar(
                r=rv + [rv[0]], theta=major_axes + [major_axes[0]],
                fill='toself', line_color='#6c3483',
                fillcolor='rgba(108,52,131,0.25)',
                name='Coverage',
            ))
            fig_radar.update_layout(
                polar=dict(radialaxis=dict(visible=True, tickfont_size=11)),
                title=dict(text='NT Axis Coverage', font_size=14),
                margin=dict(t=60,b=20),
                font=dict(size=12, color='#111'),
            )
            cr.plotly_chart(fig_radar, use_container_width=True)

        st.markdown("<div class='section-header'>Neuro-Axis Detail — Compounds &amp; Group Comparison</div>",
                    unsafe_allow_html=True)
        for axis, cnt in neuro_ctr.most_common(12):
            axis_df = neuro_df[neuro_df['Neuro_Trap'].str.contains(re.escape(axis), na=False)]
            a_up    = axis_df[axis_df['_direction'] == DIR_A]
            b_up    = axis_df[axis_df['_direction'] == DIR_B]
            n_ns    = cnt - len(a_up) - len(b_up)
            med_fc  = axis_df[fc_col].median() if fc_col in axis_df.columns else np.nan
            with st.expander(f"**{axis}** — {cnt} compounds"):
                n1, n2, n3, n4 = st.columns(4)
                n1.markdown(f"<span class='badge-adhd'>{DIR_A}</span><br>"
                            f"<b style='font-size:1.4rem;color:#c0392b'>{len(a_up)}</b>",
                            unsafe_allow_html=True)
                n2.markdown(f"<span class='badge-ctrl'>{DIR_B}</span><br>"
                            f"<b style='font-size:1.4rem;color:#1a6ea8'>{len(b_up)}</b>",
                            unsafe_allow_html=True)
                n3.markdown(f"<span style='background:#888;color:white;padding:3px 8px;border-radius:12px;"
                            f"font-size:0.85rem;font-weight:600'>n/s</span>"
                            f"<br><b style='font-size:1.4rem;color:#555'>{n_ns}</b>",
                            unsafe_allow_html=True)
                if np.isfinite(med_fc):
                    n4.metric("Median Log₂FC", f"{med_fc:+.2f}",
                              help=f"Positive = higher in {GROUP_B}")
                ea, ec = st.columns(2)
                ea.markdown(f"<span class='badge-adhd'>{DIR_A} compounds</span>",
                            unsafe_allow_html=True)
                ea.write(list(a_up[name_col].dropna().unique()))
                ec.markdown(f"<span class='badge-ctrl'>{DIR_B} compounds</span>",
                            unsafe_allow_html=True)
                ec.write(list(b_up[name_col].dropna().unique()))

    # ═════════════════════════════════════════════════════════════════════════
    # TAB 6 — GROUP A vs GROUP B (differential analysis)
    # ═════════════════════════════════════════════════════════════════════════
    with tabs[5]:
        st.markdown(f"### Differential Analysis — {GROUP_A} vs {GROUP_B}")

        vol_df = df[[fc_col, pval_col, adjp_col, 'Original_Name', name_col]].copy()
        vol_df = vol_df.dropna(subset=[fc_col, pval_col])
        vol_df['-log10p'] = -np.log10(vol_df[pval_col].clip(lower=1e-20))
        p_s = (vol_df[adjp_col] if show_adj else vol_df[pval_col])
        vol_df['Significance'] = np.where(
            (vol_df[fc_col].abs() >= fc_thresh) & (p_s < pval_thresh),
            np.where(vol_df[fc_col] > 0, DIR_B, DIR_A),
            'n/s')

        color_map = {DIR_B: COLORS[GROUP_B], DIR_A: COLORS[GROUP_A], 'n/s': '#cccccc'}
        sig_counts = vol_df['Significance'].value_counts()

        m1, m2, m3 = st.columns(3)
        m1.metric(f"↑ in {GROUP_A}",     sig_counts.get(DIR_A, 0))
        m2.metric(f"↑ in {GROUP_B}",     sig_counts.get(DIR_B, 0))
        m3.metric("Not significant",      sig_counts.get('n/s', 0))

        # Volcano
        fig_vol = px.scatter(
            vol_df, x=fc_col, y='-log10p',
            color='Significance', color_discrete_map=color_map,
            hover_data={name_col: True, fc_col: ':.2f',
                        '-log10p': ':.2f', pval_col: ':.4f', 'Significance': False},
            opacity=0.75,
            title=f'Volcano Plot  |  |FC| ≥ {fc_thresh}  |  p < {pval_thresh}'
                  f'{"  (FDR adjusted)" if show_adj else ""}',
            labels={fc_col: f'Log₂ Fold Change ({GROUP_B}/{GROUP_A})', '-log10p': '–log₁₀(p)'},
        )
        fig_vol.add_vline(x= fc_thresh, line_dash='dash', line_color='#555', line_width=1)
        fig_vol.add_vline(x=-fc_thresh, line_dash='dash', line_color='#555', line_width=1)
        fig_vol.add_hline(y=-np.log10(pval_thresh), line_dash='dash',
                          line_color='#555', line_width=1)
        # Region labels
        xmax = vol_df[fc_col].abs().max() * 0.9
        fig_vol.add_annotation(x=-xmax, y=vol_df['-log10p'].max()*0.95,
                                text=f"← Elevated in {GROUP_A}", showarrow=False,
                                font=dict(size=13, color=COLORS[GROUP_A]))
        fig_vol.add_annotation(x= xmax, y=vol_df['-log10p'].max()*0.95,
                                text=f"Elevated in {GROUP_B} →", showarrow=False,
                                font=dict(size=13, color=COLORS[GROUP_B]))
        fig_vol.update_traces(marker_size=5)
        fig_vol.update_layout(height=540, margin=dict(t=60,b=20),
                               paper_bgcolor='white', plot_bgcolor='#fafafa',
                               font=dict(size=13, color='#111'))
        st.plotly_chart(fig_vol, use_container_width=True)

        # ── Significant compound tables with controls ──────────────────────
        sig_df = (vol_df[vol_df['Significance'] != 'n/s']
                  .merge(df[['Original_Name','Main_Categories','LM_Main_Class',
                              'KEGG_Pathways','Neuro_Trap']], on='Original_Name', how='left'))

        st.markdown("<div class='section-header'>Significant Compounds — Tables</div>",
                    unsafe_allow_html=True)
        tl, tr = st.columns(2)

        with tl:
            st.markdown(f"#### ↑ Higher in {GROUP_A}")
            up_a = sig_df[sig_df['Significance'] == DIR_A].sort_values(fc_col, ascending=True)
            if len(up_a) == 0:
                st.info(f"No compounds elevated in {GROUP_A} with current filters.")
            else:
                n_max_a  = len(up_a)
                n_show_a = st.slider(f"Show top N ({GROUP_A}-elevated)", 1, n_max_a,
                                      min(25, n_max_a), key='n_adhd') if n_max_a > 1 else n_max_a
                min_fc_a = st.slider(f"|FC| minimum ({GROUP_A})", 0.0, 5.0, 0.0, 0.25, key='fc_adhd')
                shown_a  = up_a[up_a[fc_col].abs() >= min_fc_a].head(n_show_a)
                st.dataframe(shown_a[[name_col, fc_col, pval_col, 'LM_Main_Class','KEGG_Pathways']]
                             .rename(columns={name_col:'Compound', fc_col:'Log2FC', pval_col:'p-value'})
                             .reset_index(drop=True), use_container_width=True, height=350)

        with tr:
            st.markdown(f"#### ↑ Higher in {GROUP_B}")
            up_b = sig_df[sig_df['Significance'] == DIR_B].sort_values(fc_col, ascending=False)
            if len(up_b) == 0:
                st.info(f"No compounds elevated in {GROUP_B} with current filters.")
            else:
                n_max_b  = len(up_b)
                n_show_b = st.slider(f"Show top N ({GROUP_B}-elevated)", 1, n_max_b,
                                      min(25, n_max_b), key='n_ctrl') if n_max_b > 1 else n_max_b
                min_fc_b = st.slider(f"|FC| minimum ({GROUP_B})", 0.0, 5.0, 0.0, 0.25, key='fc_ctrl')
                shown_b  = up_b[up_b[fc_col].abs() >= min_fc_b].head(n_show_b)
                st.dataframe(shown_b[[name_col, fc_col, pval_col, 'LM_Main_Class','KEGG_Pathways']]
                             .rename(columns={name_col:'Compound', fc_col:'Log2FC', pval_col:'p-value'})
                             .reset_index(drop=True), use_container_width=True, height=350)

        # ── Heatmap ───────────────────────────────────────────────────────
        st.markdown("<div class='section-header'>Intensity Heatmap — Top Significant Compounds (log₁₀)</div>",
                    unsafe_allow_html=True)
        n_sig_total = int((vol_df['Significance'] != 'n/s').sum())
        n_heat_max  = max(n_sig_total, 10)
        n_heat = st.slider("Compounds in heatmap", 1, n_heat_max, min(40, n_heat_max), 5) if n_heat_max > 1 else n_heat_max
        top_sig_names = (vol_df[vol_df['Significance'] != 'n/s']
                         .nlargest(n_heat, '-log10p')['Original_Name'].tolist())
        heat_df = df[df['Original_Name'].isin(top_sig_names)].drop_duplicates('Original_Name').copy()
        if not heat_df.empty:
            sorted_samp = adhd_cols + extra_cols + control_cols
            sorted_idx  = [sample_cols.index(s) for s in sorted_samp if s in sample_cols]
            heat_vals   = heat_df[sample_cols].apply(pd.to_numeric, errors='coerce').values
            heat_log    = np.log10(np.clip(heat_vals, 1, None))[:,sorted_idx]
            labels_x    = [f"{s} ({grp_norm.get(s,'?')})" for s in sorted_samp if s in sample_cols]
            labels_y    = heat_df[name_col].fillna(heat_df['Original_Name']).tolist()
            fig_heat    = go.Figure(go.Heatmap(
                z=heat_log, x=labels_x, y=labels_y,
                colorscale='RdBu_r', zmid=float(np.nanmedian(heat_log)),
                hovertemplate='%{y}<br>%{x}<br>log₁₀ intensity: %{z:.2f}<extra></extra>',
                colorbar=dict(title=dict(text='log₁₀(intensity)', font=dict(size=12))),
            ))
            # shade each group band (A=red, extra=grey, B=blue)
            _n_a, _n_e, _n_b = len(adhd_cols), len(extra_cols), len(control_cols)
            for _x0, _x1, _fc in [
                (-0.5, _n_a - 0.5,            "rgba(192,57,43,0.07)"),
                (_n_a - 0.5, _n_a+_n_e - 0.5, "rgba(150,150,150,0.07)"),
                (_n_a+_n_e - 0.5, _n_a+_n_e+_n_b - 0.5, "rgba(26,110,168,0.07)"),
            ]:
                if _x0 < _x1:
                    fig_heat.add_vrect(x0=_x0, x1=_x1, fillcolor=_fc, line_width=0)
            fig_heat.update_layout(
                height=max(500, n_heat * 17),
                margin=dict(t=50,b=50),
                xaxis=dict(tickangle=45, tickfont_size=10),
                yaxis=dict(tickfont_size=11),
                title=dict(text=f'Top {n_heat} significant compounds | {GROUP_A} → {GROUP_B}',
                           font_size=14),
                font=dict(size=12, color='#111'),
                paper_bgcolor='white',
            )
            st.plotly_chart(fig_heat, use_container_width=True)

    # ═════════════════════════════════════════════════════════════════════════
    # TAB 7 — COMPOUND EXPLORER
    # ═════════════════════════════════════════════════════════════════════════
    with tabs[6]:
        st.markdown("### Interactive Compound Explorer")

        cf1, cf2, cf3 = st.columns(3)
        search_term = cf1.text_input("🔍 Search compound name", placeholder="e.g. Arachidonic")
        cat_filter  = cf2.multiselect("Category", sorted(df['Main_Categories'].dropna().unique()))
        sig_filter  = cf3.selectbox("Direction filter",
                                     ['All','Significant only', DIR_A, DIR_B])

        ex_df = df.copy()
        if search_term:
            mask = (ex_df[name_col].str.contains(search_term, case=False, na=False) |
                    ex_df['Original_Name'].str.contains(search_term, case=False, na=False))
            ex_df = ex_df[mask]
        if cat_filter:
            ex_df = ex_df[ex_df['Main_Categories'].apply(
                lambda x: any(c in str(x) for c in cat_filter) if pd.notna(x) else False)]
        if sig_filter == 'Significant only':
            ex_df = ex_df[ex_df['_direction'] != 'n/s']
        elif sig_filter in (DIR_A, DIR_B):
            ex_df = ex_df[ex_df['_direction'] == sig_filter]

        st.markdown(f"**{len(ex_df):,} compounds** match filters")

        disp_cols = [name_col,'Main_Categories','LM_Category','LM_Main_Class',
                     fc_col, pval_col,'OC_Ratio','KEGG_Pathways','Neuro_Trap','_direction']
        disp_cols = [c for c in disp_cols if c in ex_df.columns]
        show_df   = (ex_df[disp_cols]
                     .rename(columns={name_col:'Compound', fc_col:'Log2FC',
                                      pval_col:'p-value','_direction':'Direction'})
                     .reset_index(drop=True))
        st.dataframe(
            show_df, use_container_width=True, height=460,
            column_config={
                'Log2FC':   st.column_config.NumberColumn(format="%.2f"),
                'p-value':  st.column_config.NumberColumn(format="%.4f"),
                'OC_Ratio': st.column_config.ProgressColumn(min_value=0, max_value=1, format="%.2f"),
            }
        )

        # ── Compound Detail ───────────────────────────────────────────────
        st.markdown("<div class='section-header'>Compound Detail View</div>", unsafe_allow_html=True)

        all_names = ex_df[name_col].dropna().unique()
        if len(all_names) == 0:
            st.info("No compounds match the current filters.")
        else:
            # Text search to narrow selectbox
            det_search = st.text_input("Filter compound list", placeholder="type to narrow...",
                                        key='det_search')
            if det_search:
                all_names = [n for n in all_names if det_search.lower() in n.lower()]

            if len(all_names) > 0:
                sel_name = st.selectbox(
                    f"Select compound ({len(all_names):,} available)",
                    options=all_names[:500],
                    format_func=lambda n: (n[:80] + '…') if len(n) > 80 else n,
                )
                row = ex_df[ex_df[name_col] == sel_name].iloc[0]
                d1, d2, d3 = st.columns(3)
                for col, pairs in [
                    (d1, [('Formula', 'Formula_Final'), ('MW', 'MW_Final'),
                           ('KEGG ID', 'KEGG_ID_Final'), ('HMDB', 'HMDB_ID')]),
                    (d2, [('Category', 'Main_Categories'), ('LM Class', 'LM_Main_Class'),
                           ('O/C Ratio', 'OC_Ratio'), ('NOSC', 'NOSC')]),
                    (d3, [('Log2FC', fc_col), ('p-value', pval_col),
                           ('Neuro', 'Neuro_Trap'), ('Clinical Tags', 'Clinical_Tags')]),
                ]:
                    for label, key in pairs:
                        val = row.get(key, 'N/A')
                        if isinstance(val, float):
                            val = f"{val:.4f}" if label in ('p-value','Log2FC','O/C Ratio') else f"{val:.3f}"
                        col.markdown(f"**{label}:** {val}")

                if sample_cols:
                    int_vals = pd.to_numeric(row[sample_cols], errors='coerce')
                    int_df = pd.DataFrame({
                        'Sample': sample_cols,
                        'Intensity': int_vals.values,
                        'Group': [grp_norm.get(s,'Unknown') for s in sample_cols],
                    }).dropna(subset=['Intensity'])
                    int_df = int_df[int_df['Intensity'] > 0].sort_values('Group')
                    fig_int = px.bar(int_df, x='Sample', y='Intensity', color='Group',
                                     color_discrete_map=COLORS, log_y=True,
                                     title=f"Per-sample intensity — {sel_name[:60]}",
                                     labels={'Intensity':'Intensity (log scale)'})
                    fig_int.update_layout(xaxis_tickangle=45, margin=dict(t=60,b=90))
                    apply_template(fig_int)
                    st.plotly_chart(fig_int, use_container_width=True)

                kegg_val = row.get('KEGG_Pathways','')
                if kegg_val and str(kegg_val) != '-':
                    st.markdown(f"**KEGG Pathways:** {kegg_val}")


    # ═════════════════════════════════════════════════════════════════════════
    # TAB 8 — DOWNLOADS & EXPORTS
    # ═════════════════════════════════════════════════════════════════════════
    with tabs[7]:
        st.markdown("## 📥 Downloads & Exports")
        st.caption("All exports reflect the current filter settings (FC threshold, p-value, statistical method).")

        ecols = export_cols(df, name_col, fc_col, pval_col)
        sig = df[df['_direction'] != 'n/s'].copy()
        up_a  = df[df['_direction'] == DIR_A].copy()
        up_b  = df[df['_direction'] == DIR_B].copy()

        def dl_row(label: str, data_df: pd.DataFrame, stem: str):
            c1, c2, c3 = st.columns([3, 1, 1])
            c1.markdown(f"**{label}** — {len(data_df):,} compounds")
            c2.download_button(
                "⬇ CSV", to_csv(data_df[ecols]),
                file_name=f"{stem}.csv", mime="text/csv",
                key=f"csv_{stem}", use_container_width=True,
            )
            c3.download_button(
                "⬇ Excel", to_excel(data_df[ecols]),
                file_name=f"{stem}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"xl_{stem}", use_container_width=True,
            )

        # ── Core datasets ──────────────────────────────────────────────────
        st.markdown("### Core datasets")
        dl_row("All annotated compounds", df, "all_compounds")
        dl_row(f"Significant compounds (p<{pval_thresh}, |FC|≥{fc_thresh})", sig, "significant_compounds")
        dl_row(f"↑ {GROUP_A} compounds", up_a, f"up_{GROUP_A.lower()}")
        dl_row(f"↑ {GROUP_B} compounds", up_b, f"up_{GROUP_B.lower()}")

        st.divider()

        # ── Thematic slices ────────────────────────────────────────────────
        st.markdown("### Thematic slices")

        lipid_df = sig[sig.get('Main_Categories', pd.Series(dtype=str)).astype(str).str.contains('Lipid|lipid', na=False)] \
            if 'Main_Categories' in sig.columns else sig.iloc[0:0]
        kegg_df  = sig[sig['KEGG_Pathways'].astype(str).str.strip().str.replace('-','').str.len() > 0] \
            if 'KEGG_Pathways' in sig.columns else sig.iloc[0:0]
        neuro_df = sig[sig['Neuro_Trap'].astype(str).str.strip().str.lower().isin(['yes','true','1'])] \
            if 'Neuro_Trap' in sig.columns else sig.iloc[0:0]
        hoc_df   = sig[pd.to_numeric(sig.get('OC_Ratio', pd.Series(dtype=float)), errors='coerce') > 0.4] \
            if 'OC_Ratio' in sig.columns else sig.iloc[0:0]

        dl_row("Significant lipids", lipid_df, "sig_lipids")
        dl_row("Significant KEGG-mapped compounds", kegg_df, "sig_kegg_mapped")
        dl_row("Neuro-active significant compounds", neuro_df, "sig_neuroactive")
        dl_row("High-oxidation compounds (O/C > 0.4)", hoc_df, "sig_high_oxidation")

        st.divider()

        # ── Full data with sample intensities ─────────────────────────────
        st.markdown("### Full data with sample intensities")
        if sample_cols:
            int_cols = [c for c in ([name_col, fc_col, pval_col, '_direction']
                                    + [c2 for c2 in EXPORT_BASE if c2 in df.columns and c2 != name_col]
                                    + sample_cols) if c in df.columns]
            dl_row("All compounds + raw intensities", df[int_cols], "all_with_intensities")
            dl_row("Significant compounds + raw intensities", sig[int_cols], "sig_with_intensities")
        else:
            st.info("No sample intensity columns detected in the current dataset.")

        st.divider()

        # ── MetaboAnalyst export ───────────────────────────────────────────
        st.markdown("### MetaboAnalyst Export")
        st.markdown(
            "Two files required by [MetaboAnalyst](https://www.metaboanalyst.ca/) "
            "for statistical analysis and pathway enrichment."
        )

        if not sample_cols:
            st.warning("Sample intensity columns not detected — MetaboAnalyst export unavailable.")
        else:
            peak_table, meta_table, n_inc, n_exc = build_metaboanalyst(df, sample_cols, grp_norm)

            col_pt, col_mt = st.columns(2)
            _hdr = 'style="min-height:52px;line-height:1.5;margin-bottom:4px"'
            with col_pt:
                warn = (f'<br><small style="color:#888">⚠ {n_exc} compounds excluded '
                        f'(no recognized identifier).</small>') if n_exc else ''
                st.markdown(
                    f'<div {_hdr}><b>Peak Table</b> — {n_inc:,} compounds '
                    f'with PubChem / HMDB / KEGG ID{warn}</div>',
                    unsafe_allow_html=True,
                )
                st.dataframe(peak_table.head(5), use_container_width=True, height=200)
                st.download_button(
                    "⬇ Peak Table CSV (MetaboAnalyst)",
                    to_csv(peak_table),
                    file_name="metaboanalyst_peak_table.csv",
                    mime="text/csv",
                    key="ma_peak",
                    use_container_width=True,
                )

            with col_mt:
                st.markdown(
                    f'<div {_hdr}><b>Sample Metadata</b> — '
                    f'{len(meta_table)} samples × 2 columns</div>',
                    unsafe_allow_html=True,
                )
                st.dataframe(meta_table, use_container_width=True, height=200)
                st.download_button(
                    "⬇ Sample Metadata CSV (MetaboAnalyst)",
                    to_csv(meta_table),
                    file_name="metaboanalyst_metadata.csv",
                    mime="text/csv",
                    key="ma_meta",
                    use_container_width=True,
                )

            with st.expander("How to use in MetaboAnalyst"):
                st.markdown(
                    f"""
1. Go to **metaboanalyst.ca → Statistical Analysis → Two-group Comparison**
2. Upload **Peak Table CSV** as the data file (rows = compounds, columns = samples)
3. The first column (`PubChem_CID`) contains the identifier — select **PubChem CID** as ID type
4. Upload **Sample Metadata CSV** for group labels — column `Sample` maps to sample names, `Group` is the class label
5. For pathway analysis, use **Pathway Analysis** module and upload the same Peak Table
"""
                )


# ─── Entry point ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    main()
